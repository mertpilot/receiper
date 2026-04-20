from pathlib import Path

from openpyxl import Workbook, load_workbook


TEMPLATE_HEADERS = [
    "Kod",
    "Hesap Kodu",
    "Evrak Tarihi",
    "Evrak No",
    "Vergi / TC No",
    "Gider Aciklama",
    "KDV%",
    "Alinan Mal/Masraf",
    "Ind.KDV",
    "Toplam",
]

HEADERS = TEMPLATE_HEADERS


def _round2(value):
    if value is None:
        return None
    return round(float(value), 2)


def build_template_row(parsed: dict) -> dict:
    total = _round2(parsed.get("total"))
    vat_amount = _round2(parsed.get("kdv"))
    net_amount = _round2(parsed.get("net_amount"))

    if net_amount is None and total is not None and vat_amount is not None:
        net_amount = _round2(max(total - vat_amount, 0))
    if net_amount is None:
        net_amount = total

    description = (parsed.get("expense_description") or "").strip()
    if not description:
        merchant = (parsed.get("merchant") or "").strip()
        description = f"{merchant} Gideri".strip() if merchant else "Fis Gideri"

    return {
        "kod": "MS",
        "hesap_kodu": "MS",
        "evrak_tarihi": parsed.get("date", ""),
        "evrak_no": parsed.get("receipt_no", ""),
        "vergi_tc_no": parsed.get("tax_id", ""),
        "gider_aciklama": description,
        "kdv_orani": _round2(parsed.get("vat_rate")),
        "alinan_mal_masraf": net_amount,
        "ind_kdv": vat_amount,
        "toplam": total,
    }


def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "GiderSayfasi"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    wb.save(path)


def _template_row_values(template_row: dict) -> list:
    return [
        template_row["kod"],
        template_row["hesap_kodu"],
        template_row["evrak_tarihi"],
        template_row["evrak_no"],
        template_row["vergi_tc_no"],
        template_row["gider_aciklama"],
        template_row["kdv_orani"],
        template_row["alinan_mal_masraf"],
        template_row["ind_kdv"],
        template_row["toplam"],
    ]


def _append_with_openpyxl(path: Path, row_values: list) -> int:
    wb = load_workbook(path)
    ws = wb.active
    ws.append(row_values)
    row_number = ws.max_row
    wb.save(path)
    return row_number


def _append_with_excel_com(path: Path, row_values: list) -> int | None:
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception:
        return None

    app = None
    workbook = None
    opened_here = False
    created_app = False

    try:
        pythoncom.CoInitialize()
        target = str(path.resolve()).lower()

        try:
            app = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            app = win32com.client.Dispatch("Excel.Application")
            created_app = True

        app.DisplayAlerts = False

        for wb in app.Workbooks:
            if str(Path(wb.FullName).resolve()).lower() == target:
                workbook = wb
                break

        if workbook is None:
            workbook = app.Workbooks.Open(str(path.resolve()))
            opened_here = True

        sheet = workbook.Worksheets(1)
        xl_up = -4162
        last_row = int(sheet.Cells(sheet.Rows.Count, 1).End(xl_up).Row)
        row_number = max(1, last_row) + 1

        for col, value in enumerate(row_values, start=1):
            sheet.Cells(row_number, col).Value = "" if value is None else value

        workbook.Save()
        if opened_here:
            workbook.Close(SaveChanges=True)
        if created_app and app.Workbooks.Count == 0:
            app.Quit()
        return row_number
    except Exception:
        return None
    finally:
        try:
            pythoncom.CoUninitialize()  # type: ignore[name-defined]
        except Exception:
            pass


def append_receipt_row(path: Path, parsed: dict, file_name: str, raw_text: str) -> int:
    ensure_workbook(path)
    template_row = build_template_row(parsed)
    row_values = _template_row_values(template_row)

    try:
        return _append_with_openpyxl(path, row_values)
    except PermissionError as openpyxl_error:
        row = _append_with_excel_com(path, row_values)
        if row is not None:
            return row
        raise openpyxl_error
